import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from core.api.schemas.subdivision import SubdivisionBase, SubdivisionResponse
from core.api.schemas.position import PositionRespone
from core.api.tools.subdivision_tools import get_all_subdivisions_from_db, add_subdivision, get_subdivision_by_id_from_db
from core.models.database import get_async_session
from core.api.tools.position_tools import get_all_positions_from_db, add_position_in_db
from core.api.tools.employee_tools import add_import_employee, get_employee_from_db
from core.api.tools.division_tools import get_all_divisions_from_db
from core.api.tools.exam_tools import get_exams_for_next_month_from_db
from sqlalchemy.ext.asyncio import AsyncSession
from core.api.tools.employee_tools import get_employee_from_db
from core.api.schemas.position import CreatePosition
from core.models.employee import Division, Subdivision, Employee
from core.CustomExceptions import CustomException
from dateutil.relativedelta import relativedelta
from core.models.User import User
from core.UserManager import current_active_verified_user
import sys
import io

from core.models.exam import Exam

router = APIRouter(
    prefix="/file",
    tags=["file"]
)



@router.post("/upload")
async def upload_file(file: UploadFile, 
                      user: User = Depends(current_active_verified_user),
                      session: AsyncSession = Depends(get_async_session)):
    try:
        contents = file.file.read()
        filename=io.BytesIO(contents)
        book = load_workbook(filename=filename)
        sheet = book.active
        i = 2
        employee_list = [] 
        while sheet[f'A{i}'].value:
            fio = sheet[f'A{i}'].value
            subdivision = sheet[f'B{i}'].value
            position = sheet[f'C{i}'].value
            employee_list.append(EmployeeInfo(fio=fio, position=position, subdivision=subdivision))
            i += 1
        if i == 2:
            raise CustomException("Excel файл пустой")
        subdivisions: list[SubdivisionBase] = await get_all_subdivisions_from_db(session=session)
        positions: list[PositionRespone] = await get_all_positions_from_db(session=session)
        for employee in employee_list:
            for sub in subdivisions:
                if employee.subdivision[0] == sub.name:
                    employee.subdivision[1] = sub.id
                    break
            for pos in positions:
                if employee.position[0] == pos.name:
                    employee.position[1] = pos.id
                    break
            if not employee.position[1]:
                new_position = await add_position_in_db(new_position=CreatePosition(name=employee.position[0]), session=session)
                positions.append(PositionRespone(name=new_position.name, id=new_position.id))
                employee.position[1] = new_position.id
            
                
            employee_tmp = await get_employee_from_db(session=session, name=employee.fio, id=None)
            if not employee_tmp:
                await add_import_employee(fio=employee.fio,
                                          position_id=employee.position[1],
                                          subdivision_id=employee.subdivision[1],
                                          session=session)
        return {"details": 'succes'}
    except:
        type, value, traceback = sys.exc_info()
        raise HTTPException(status_code=400, detail=str(value))


@router.get("/get-month-exam-list")
async def download_file(user: User = Depends(current_active_verified_user),
                        session: AsyncSession = Depends(get_async_session)):
    try:
        exams: list[Exam] = await get_exams_for_next_month_from_db(session=session)
        divisions: list[Division] = await get_all_divisions_from_db(session=session)
        wb = Workbook()
        ws = wb.active

        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"))
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

        ws.cell(1, 1).value = 'ФИО'
        ws.cell(1, 2).value = 'Экзамен'
        ws.cell(1, 3).value = 'Дата сдачи'
        ws.cell(1, 4).value = 'Примечание'
        for j in range(4):
            ws.cell(1, j+1).border = border
            ws.cell(1, j+1).alignment = Alignment(horizontal="center")
        i = 2
        for division in divisions:
            ws.cell(i, 1).value = f'Подразделение: {division.name}'
           
            i += 1
            for exam in exams:
                res: Employee = await get_employee_from_db(session=session, name=None, id=exam.employee_id)
                res: Subdivision = await get_subdivision_by_id_from_db(subdivision_id=res.subdivision_id, session=session)
                if res.division_id == division.id:
                    ws.cell(i, 1).value = exam.employee.FIO
                    ws.cell(i, 2).value = exam.exam_type.name
                    ws.cell(i, 3).value = exam.next_date
                    ws.cell(i, 4).value = exam.notation
                
                    for j in range(4):
                        ws.cell(i, j+1).border = border
                        if j >= 1:
                            ws.cell(i, j+1).alignment = Alignment(horizontal="center")
                    i += 1
            i += 1
        today = datetime.date.today()
        next_month = today.replace(day=1) + relativedelta(months=1)
        file_name = f'exams-{next_month.year}-{next_month.month}.xlsx'
        file_path = f'static/documents/{file_name}'
        wb.save(file_path)
        return FileResponse(path=file_path, filename=file_name, media_type='multipart/form-data')
    except:
        type, value, traceback = sys.exc_info()
        raise HTTPException(status_code=400, detail=str(value))
    
class EmployeeInfo:
    def __init__(self, fio: str, position: str, subdivision: str):
        self.fio = fio
        self.position = [position, None]
        self.subdivision = [subdivision, None]
